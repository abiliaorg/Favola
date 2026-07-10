# -*- coding: utf-8 -*-
"""Scoring MT dei test di comprensione e correlazione con la modalità (text/images).

- Chiave: study/chiave_derivata.json (fogli adattati, validata sulle prove MT ufficiali).
- Punteggio grezzo MT = numero risposte corrette; bianche e doppie = non corrette.
- Fascia MT (CCRD/PSD/RAD/RIDI) con le soglie ufficiali; nota: prove applicate per
  tutti all'uscita, fasce da intendersi come riferimento descrittivo.
- Modalità: dal gaze quando esiste, altrimenti dal disegno (track B: 1a storia text,
  2a images; track F: inverso), marcata come "disegno".

Output: study/risultati_test.xlsx (fogli Punteggi + Riepilogo) + statistiche a video.
"""
import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

STUDY = Path(__file__).resolve().parent.parent
GAZE = STUDY.parent / "data" / "02_gaze"

data = json.loads((STUDY / "chiave_derivata.json").read_text(encoding="utf-8"))
KEY = {k: v for k, v in data.items() if not k.startswith("_")}
FASCE = data["_fasce_mt"]
ALT = data.get("_accetta_anche", {})  # es. orso Q11: accetta anche 'c' (stimolo dice "pesce")

def is_correct(storia, qi, given):
    """qi 1-based; True se la risposta è la chiave o un'alternativa accettata."""
    return given == KEY[storia][qi - 1] or given in ALT.get(storia, {}).get(str(qi), [])

_map = json.loads((STUDY / "mappatura_soggetti.json").read_text(encoding="utf-8"))
MAPPING, EXCLUDED = _map["mapping"], _map["excluded"]

# demografia (ID, AGE, SEX) — eventuali righe duplicate vengono ignorate
DEMO = {}
import csv
with open(STUDY / "demo.csv", newline="", encoding="utf-8-sig") as fh:
    for row in csv.DictReader(fh, delimiter="\t"):
        pid = row["ID"].strip().upper()
        if pid in DEMO and (DEMO[pid]["age"] != int(row["AGE"]) or DEMO[pid]["sex"] != row["SEX"].strip().upper()):
            print(f"ATTENZIONE: demo.csv ha righe contrastanti per {pid}")
        DEMO[pid] = {"age": int(row["AGE"]), "sex": row["SEX"].strip().upper()}

EN2IT = {"fox": "volpe e boscaiolo", "carpet": "tappeto", "cats": "gatta", "yawn": "sbadiglio",
         "dolphin": "delfino", "panda": "panda", "bear": "orso", "eels": "anguille"}
FIRST_STORIES = {"volpe e boscaiolo", "gatta", "delfino", "anguille"}

def fascia(storia, correct):
    f = FASCE[storia]
    if correct >= f["CCRD"]:
        return "CCRD"
    if correct >= f["PSD"]:
        return "PSD"
    if correct >= f["RAD"]:
        return "RAD"
    return "RIDI"

# --- modalità effettiva dalle registrazioni gaze ---
gaze_mod, gaze_cls = {}, {}
for folder in sorted(GAZE.iterdir()):
    if not folder.is_dir() or not folder.name.startswith("data_"):
        continue
    for f in sorted(folder.glob("*.json")):
        d = json.loads(f.read_text(encoding="utf-8"))
        pid = str(d.get("participantId", "")).strip("`").strip().upper()
        if not re.match(r"^[BF]\d+$", pid):
            continue
        st = EN2IT.get(d.get("story"))
        if st:
            gaze_mod[(pid, st)] = d.get("typology")
            gaze_cls[pid] = str(d.get("class", ""))

def design_modality(pid, storia):
    first = storia in FIRST_STORIES
    if pid[0] == "B":
        return "text" if first else "images"
    return "images" if first else "text"

# --- risposte dei bambini ---
wb = openpyxl.load_workbook(STUDY / "test.xlsx")
ws = wb["PROVE"]
rows = list(ws.iter_rows(values_only=True))

records, skipped = [], []
for r in rows[1:]:
    pid_raw, storia = r[0], r[1]
    if not pid_raw or not storia:
        continue
    pid = str(pid_raw).strip().upper()
    storia = str(storia).strip()
    key = KEY[storia]
    given = []
    for i in range(len(key)):
        a = r[2 + i]
        given.append(str(a).strip().lower() if a is not None else "")
    if all(a == "" for a in given):
        skipped.append((pid, storia))
        continue
    correct = sum(1 for i, a in enumerate(given, start=1) if is_correct(storia, i, a))
    blank = sum(1 for a in given if a == "")
    invalid = sum(1 for a in given if a and (len(a) > 1 or a not in "abcd"))
    n = len(key)
    mod = gaze_mod.get((pid, storia))
    mod_src = "gaze"
    if mod is None:
        mod, mod_src = design_modality(pid, storia), "disegno"
    records.append({
        "pid": pid, "newid": MAPPING.get(pid),
        "escluso": None if pid in MAPPING else EXCLUDED.get(pid, "non in mappatura"),
        "age": DEMO.get(pid, {}).get("age"), "sex": DEMO.get(pid, {}).get("sex"),
        "track": pid[0], "classe": gaze_cls.get(pid, ""),
        "storia": storia, "mod": mod, "mod_src": mod_src,
        "pos": 1 if storia in FIRST_STORIES else 2,
        "n": n, "correct": correct, "wrong": n - correct - blank, "blank": blank,
        "invalid": invalid, "pct": round(correct / n * 100, 1),
        "fascia": fascia(storia, correct),
    })

# --- solo soggetti validi (mappati TI/IT); esclusi documentati a parte ---
excluded_records = [r for r in records if r["escluso"]]
records = [r for r in records if not r["escluso"]]

# --- excel ---
out = openpyxl.Workbook()
ws1 = out.active
ws1.title = "Punteggi"
hdr = ["ID nuovo", "ID orig.", "Gruppo", "Classe", "Età", "Sesso", "Storia", "Modalità",
       "Fonte modalità", "Posizione", "N domande", "Corrette", "Sbagliate", "In bianco",
       "Doppie/invalide", "Punteggio %", "Fascia MT"]
ws1.append(hdr)
FILL = {"CCRD": "C6EFCE", "PSD": "DDEBF7", "RAD": "FFEB9C", "RIDI": "FFC7CE"}
for c in ws1[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4472C4")
for r in sorted(records, key=lambda r: (r["newid"][:2], int(r["newid"][2:]), r["pos"])):
    ws1.append([r["newid"], r["pid"], r["newid"][:2], r["classe"], r["age"], r["sex"],
                r["storia"], r["mod"], r["mod_src"], r["pos"], r["n"], r["correct"],
                r["wrong"], r["blank"], r["invalid"], r["pct"], r["fascia"]])
    ws1.cell(ws1.max_row, 17).fill = PatternFill("solid", fgColor=FILL[r["fascia"]])
for col, w in zip("ABCDEFGHIJKLMNOPQ", [9, 8, 8, 8, 6, 7, 18, 10, 14, 10, 11, 9, 10, 10, 14, 12, 10]):
    ws1.column_dimensions[col].width = w
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:Q{ws1.max_row}"

wsE = out.create_sheet("Esclusi")
wsE.append(["ID orig.", "Storia", "Punteggio %", "Motivo esclusione"])
for c in wsE[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="C00000")
for r in sorted(excluded_records, key=lambda r: (r["pid"][0], int(re.sub(r"\D", "", r["pid"]) or 0))):
    wsE.append([r["pid"], r["storia"], r["pct"], r["escluso"]])
for col, w in zip("ABCD", [9, 18, 12, 44]):
    wsE.column_dimensions[col].width = w

def mean(xs):
    return sum(xs) / len(xs) if xs else float("nan")
def sd(xs):
    if len(xs) < 2:
        return float("nan")
    m = mean(xs)
    return (sum((x - m) ** 2 for x in xs) / (len(xs) - 1)) ** 0.5
def tstat(xs):
    return mean(xs) / (sd(xs) / len(xs) ** 0.5) if len(xs) > 1 and sd(xs) > 0 else float("nan")

ws2 = out.create_sheet("Riepilogo")
def w2(*vals, bold=False):
    ws2.append(list(vals))
    if bold:
        for c in ws2[ws2.max_row]:
            c.font = Font(bold=True)

lines = []
def report(s=""):
    print(s)
    lines.append(s)

report(f"Test validi: {len(records)}; righe vuote saltate: {skipped}")

w2("PUNTEGGIO PER MODALITÀ (tutti i test)", bold=True)
w2("Modalità", "n", "media %", "sd")
report("\n=== Punteggio medio per modalità ===")
for m in ["text", "images"]:
    xs = [r["pct"] for r in records if r["mod"] == m]
    w2(m, len(xs), round(mean(xs), 1), round(sd(xs), 1))
    report(f"  {m:6}: n={len(xs):3}  media={mean(xs):5.1f}%  sd={sd(xs):4.1f}")

report("\n=== Confronto appaiato (1 text + 1 images per bambino) ===")
bypid = {}
for r in records:
    bypid.setdefault(r["pid"], []).append(r)
pairs, excluded = [], []
for pid, rs in sorted(bypid.items()):
    if len(rs) == 2 and sorted(x["mod"] for x in rs) == ["images", "text"]:
        si = next(x["pct"] for x in rs if x["mod"] == "images")
        st = next(x["pct"] for x in rs if x["mod"] == "text")
        pairs.append((pid, si - st, si, st))
    else:
        excluded.append((pid, [f"{x['storia'][:12]}/{x['mod']}" for x in rs]))
d = [x[1] for x in pairs]
t = mean(d) / (sd(d) / len(d) ** 0.5) if len(d) > 1 and sd(d) > 0 else float("nan")
report(f"  n coppie={len(d)}  diff media (images-text)={mean(d):+.1f} pt  sd={sd(d):.1f}  t={t:.2f} (df={len(d)-1})")
report(f"  images meglio: {sum(1 for x in d if x > 0)}, text meglio: {sum(1 for x in d if x < 0)}, pari: {sum(1 for x in d if x == 0)}")
report("  esclusi: " + ("; ".join(f"{p} ({', '.join(v)})" for p, v in excluded) or "nessuno"))
d_noB5 = [x[1] for x in pairs if x[0] != "B5"]
report(f"  sensibilità senza B5: n={len(d_noB5)} diff media={mean(d_noB5):+.1f} pt")

w2()
w2("CONFRONTO APPAIATO images - text", bold=True)
w2("n coppie", len(d))
w2("diff media (punti %)", round(mean(d), 1))
w2("t appaiato", round(t, 2))
w2("images meglio / text meglio / pari",
   f"{sum(1 for x in d if x > 0)} / {sum(1 for x in d if x < 0)} / {sum(1 for x in d if x == 0)}")

report("\n=== Per storia ===")
w2()
w2("PER STORIA", bold=True)
w2("Storia", "text n", "text media %", "images n", "images media %")
for st in sorted({r["storia"] for r in records}):
    cells = [st]
    line = f"  {st:18}"
    for m in ["text", "images"]:
        xs = [r["pct"] for r in records if r["storia"] == st and r["mod"] == m]
        cells += [len(xs), round(mean(xs), 1) if xs else None]
        line += f"  {m}: n={len(xs):2} media={mean(xs):5.1f}%" if xs else f"  {m}: n= 0        "
    w2(*cells)
    report(line)

report("\n=== Per classe ===")
w2()
w2("PER CLASSE", bold=True)
w2("Classe", "text n", "text media %", "images n", "images media %")
for cl in sorted({r["classe"] for r in records if r["classe"]}):
    cells = [cl]
    line = f"  classe {cl}"
    for m in ["text", "images"]:
        xs = [r["pct"] for r in records if r["classe"] == cl and r["mod"] == m]
        cells += [len(xs), round(mean(xs), 1) if xs else None]
        line += f"  {m}: n={len(xs):2} media={mean(xs):5.1f}%" if xs else f"  {m}: n= 0"
    w2(*cells)
    report(line)

report("\n=== Per gruppo (ordine reale) ===")
for grp in ["TI", "IT"]:
    xs = [r["pct"] for r in records if r["newid"].startswith(grp)]
    report(f"  {grp}: n={len(xs):3} media={mean(xs):5.1f}%  sd={sd(xs):4.1f}")

report("\n=== Fasce MT per modalità ===")
w2()
w2("FASCE MT PER MODALITÀ", bold=True)
w2("Modalità", "CCRD", "PSD", "RAD", "RIDI")
for m in ["text", "images"]:
    cnt = {f: sum(1 for r in records if r["mod"] == m and r["fascia"] == f)
           for f in ["CCRD", "PSD", "RAD", "RIDI"]}
    w2(m, cnt["CCRD"], cnt["PSD"], cnt["RAD"], cnt["RIDI"])
    tot = sum(cnt.values())
    report(f"  {m:6}: " + "  ".join(f"{f}={cnt[f]} ({cnt[f]/tot*100:.0f}%)" for f in cnt))

# --- demografia ---
kids = {}
for r in records:
    kids.setdefault(r["newid"], {"grp": r["newid"][:2], "age": r["age"], "sex": r["sex"],
                                 "pcts": []})["pcts"].append(r["pct"])
for k in kids.values():
    k["score"] = mean(k["pcts"])

report("\n=== Demografia e bilanciamento TI/IT ===")
w2()
w2("DEMOGRAFIA (soggetti validi)", bold=True)
w2("Gruppo", "n", "M", "F", "età media", "età min-max")
for grp in ["TI", "IT"]:
    ks = [k for k in kids.values() if k["grp"] == grp]
    ages = [k["age"] for k in ks if k["age"] is not None]
    nm = sum(1 for k in ks if k["sex"] == "M")
    nf = sum(1 for k in ks if k["sex"] == "F")
    w2(grp, len(ks), nm, nf, round(mean(ages), 1), f"{min(ages)}-{max(ages)}")
    report(f"  {grp}: n={len(ks)}  M={nm} F={nf}  età media={mean(ages):.1f} (range {min(ages)}-{max(ages)})")
missing_demo = sorted({r["pid"] for r in records if r["age"] is None})
if missing_demo:
    report(f"  senza dati demografici: {', '.join(missing_demo)}")

report("\n=== Punteggio per sesso (media dei 2 test per bambino) ===")
w2()
w2("PUNTEGGIO PER SESSO", bold=True)
w2("Sesso", "n", "media %", "sd")
for sx in ["M", "F"]:
    xs = [k["score"] for k in kids.values() if k["sex"] == sx]
    w2(sx, len(xs), round(mean(xs), 1), round(sd(xs), 1))
    report(f"  {sx}: n={len(xs):2}  media={mean(xs):5.1f}%  sd={sd(xs):4.1f}")
xm = [k["score"] for k in kids.values() if k["sex"] == "M"]
xf = [k["score"] for k in kids.values() if k["sex"] == "F"]
tw = (mean(xm) - mean(xf)) / (sd(xm) ** 2 / len(xm) + sd(xf) ** 2 / len(xf)) ** 0.5
report(f"  diff (M-F)={mean(xm)-mean(xf):+.1f} pt  t(Welch)={tw:.2f}")

report("\n=== Modalità per sesso (diff appaiata images-text) ===")
for sx in ["M", "F"]:
    ds = []
    for nid, k in kids.items():
        if k["sex"] != sx:
            continue
        rt = next((r["pct"] for r in records if r["newid"] == nid and r["mod"] == "text"), None)
        ri = next((r["pct"] for r in records if r["newid"] == nid and r["mod"] == "images"), None)
        if rt is not None and ri is not None:
            ds.append(ri - rt)
    report(f"  {sx}: n={len(ds):2}  diff media={mean(ds):+.1f} pt  t={tstat(ds) if len(ds)>1 else float('nan'):.2f}")

report("\n=== Età e punteggio ===")
pts = [(k["age"], k["score"]) for k in kids.values() if k["age"] is not None]
ma, ms = mean([a for a, _ in pts]), mean([s for _, s in pts])
cov = sum((a - ma) * (s - ms) for a, s in pts) / (len(pts) - 1)
r_as = cov / (sd([a for a, _ in pts]) * sd([s for _, s in pts]))
report(f"  correlazione età-punteggio (tutti, n={len(pts)}): r={r_as:+.2f}  "
       f"(NB: l'età è quasi coincidente con la classe, quindi confusa con la prova)")
for age in sorted({a for a, _ in pts}):
    xs = [s for a, s in pts if a == age]
    report(f"    {age} anni: n={len(xs):2}  media={mean(xs):5.1f}%")

w2()
w2("Nota metodologica: prove MT applicate per tutti all'uscita (fine anno); somministrazione",)
w2("non standard (ascolto + caption, non lettura), fasce da intendersi come riferimento descrittivo.")

for col, w in zip("ABCDE", [38, 12, 14, 12, 16]):
    ws2.column_dimensions[col].width = w

out.save(STUDY / "risultati_test.xlsx")
report(f"\nSalvato: {STUDY / 'risultati_test.xlsx'}")
