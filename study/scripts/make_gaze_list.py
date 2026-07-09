# -*- coding: utf-8 -*-
"""Genera un elenco Excel delle registrazioni gaze in data/02_gaze."""
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent.parent / "data" / "02_gaze"
PCS = {d.name: ("Benedetta" if d.name.startswith("data_bene") else "Francesca")
       for d in sorted(BASE.iterdir()) if d.is_dir()}
OUT = BASE / "registrazioni_gaze.xlsx"

TEST_IDS = re.compile(r"^(test\d*|testaudio|prova)$", re.IGNORECASE)

rows = []
for folder, pc in PCS.items():
    for f in sorted((BASE / folder).glob("*.json")):
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
        except Exception as e:
            rows.append({"pc": pc, "file": f.name, "errore": str(e)})
            continue
        samples = data.get("samples") or []
        duration = round(samples[-1]["t"], 1) if samples else 0
        date_iso = data.get("date", "")
        # nome file: YYYYMMDD_HHMMSS_... -> data/ora locali
        m = re.match(r"(\d{4})(\d{2})(\d{2})_(\d{2})(\d{2})(\d{2})_", f.name)
        if m:
            date_str = f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
            time_str = f"{m.group(4)}:{m.group(5)}:{m.group(6)}"
        else:
            date_str, time_str = date_iso[:10], date_iso[11:19]
        pid = str(data.get("participantId", ""))
        rows.append({
            "pc": pc,
            "cartella": folder,
            "file": f.name,
            "data": date_str,
            "ora": time_str,
            "partecipante": pid,
            "classe": data.get("class", ""),
            "storia": data.get("story", ""),
            "modalita": data.get("typology", ""),
            "durata": duration,
            "campioni": len(samples),
            "calibrazione": "sì" if data.get("calibration") else "no",
            "test": "sì" if TEST_IDS.match(pid) else "",
        })

wb = Workbook()
ws = wb.active
ws.title = "Registrazioni"

headers = ["PC", "Cartella", "File", "Data", "Ora", "Partecipante", "Classe", "Storia",
           "Modalità", "Durata (s)", "N. campioni", "Calibrazione", "Test"]
ws.append(headers)
header_fill = PatternFill("solid", fgColor="4472C4")
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

test_fill = PatternFill("solid", fgColor="FFF2CC")
for r in rows:
    ws.append([r.get("pc"), r.get("cartella"), r.get("file"), r.get("data"), r.get("ora"),
               r.get("partecipante"), r.get("classe"), r.get("storia"),
               r.get("modalita"), r.get("durata"), r.get("campioni"),
               r.get("calibrazione"), r.get("test")])
    if r.get("test") == "sì":
        for cell in ws[ws.max_row]:
            cell.fill = test_fill

widths = [12, 13, 46, 11, 10, 13, 8, 10, 10, 11, 12, 13, 6]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
ws.freeze_panes = "A2"

try:
    wb.save(OUT)
    saved = OUT
except PermissionError:
    saved = OUT.with_name(OUT.stem + "_new.xlsx")
    wb.save(saved)
    print(f"ATTENZIONE: {OUT.name} è aperto in un altro programma, salvato come {saved.name}")

n_bene = sum(1 for r in rows if r["pc"] == "Benedetta")
n_fra = sum(1 for r in rows if r["pc"] == "Francesca")
n_test = sum(1 for r in rows if r.get("test") == "sì")
print(f"Salvato: {saved}")
print(f"Totale registrazioni: {len(rows)} (Benedetta: {n_bene}, Francesca: {n_fra}, di cui test/prova: {n_test})")
