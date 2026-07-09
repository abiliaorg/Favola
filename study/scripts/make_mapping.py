# -*- coding: utf-8 -*-
"""Crea la tabella di mappatura ID originale -> ID nuovo (TI01.., IT01..).

Regole concordate:
- Si rinominano SOLO i soggetti con ordine verificato dal gaze (TI o IT certi).
- Esclusi (non rinominati e rimossi dall'analisi): ordine non verificabile
  (F1, F2, F3, F6, F10, F25), somministrazione anomala (F4 = images+images,
  F35 = text+text), assenti (B19).
- Numerazione progressiva per gruppo nell'ordine degli ID originali.

Output: study/mappatura_soggetti.xlsx + study/mappatura_soggetti.json
"""
import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

STUDY = Path(__file__).resolve().parent.parent
GAZE = STUDY.parent / "data" / "02_gaze"

EN2IT = {"fox": "volpe e boscaiolo", "carpet": "tappeto", "cats": "gatta", "yawn": "sbadiglio",
         "dolphin": "delfino", "panda": "panda", "bear": "orso", "eels": "anguille"}

recs = {}
for folder in sorted(GAZE.iterdir()):
    if not folder.is_dir() or not folder.name.startswith("data_"):
        continue
    for f in sorted(folder.glob("*.json")):
        d = json.loads(f.read_text(encoding="utf-8"))
        pid = str(d.get("participantId", "")).strip("`").strip().upper()
        if not re.match(r"^[BF]\d+$", pid) or d.get("story") not in EN2IT:
            continue
        recs.setdefault(pid, []).append({
            "mod": d.get("typology"), "time": f.name[:15],
            "classe": str(d.get("class", "")),
        })

def sortkey(pid):
    return (pid[0], int(re.sub(r"\D", "", pid)))

order_of = {}
classe_of = {}
for pid, rr in recs.items():
    rr = sorted(rr, key=lambda x: x["time"])
    classe_of[pid] = rr[0]["classe"]
    if len(rr) == 2:
        seq = "".join("T" if x["mod"] == "text" else "I" for x in rr)
        if seq in ("TI", "IT"):
            order_of[pid] = seq

EXCLUDED = {
    "F1": "nessun gaze: ordine non verificabile", "F6": "nessun gaze: ordine non verificabile",
    "F2": "gaze parziale: ordine non verificabile", "F3": "gaze parziale: ordine non verificabile",
    "F10": "gaze parziale: ordine non verificabile", "F25": "gaze parziale: ordine non verificabile",
    "F4": "somministrazione anomala (images+images)",
    "F35": "somministrazione anomala (text+text)",
    "B19": "assente (nessun dato)",
}

mapping = {}
for grp in ("TI", "IT"):
    pids = sorted([p for p, o in order_of.items() if o == grp], key=sortkey)
    for i, pid in enumerate(pids, start=1):
        mapping[pid] = f"{grp}{i:02d}"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mappatura"
ws.append(["ID originale", "ID nuovo", "Gruppo", "Classe"])
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4472C4")
for pid in sorted(mapping, key=sortkey):
    ws.append([pid, mapping[pid], mapping[pid][:2], classe_of.get(pid, "")])
ws2 = wb.create_sheet("Esclusi")
ws2.append(["ID originale", "Motivo esclusione"])
for c in ws2[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="C00000")
for pid in sorted(EXCLUDED, key=sortkey):
    ws2.append([pid, EXCLUDED[pid]])
for w, col in zip([14, 10, 8, 8], "ABCD"):
    ws.column_dimensions[col].width = w
ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 46
wb.save(STUDY / "mappatura_soggetti.xlsx")

(STUDY / "mappatura_soggetti.json").write_text(
    json.dumps({"mapping": mapping, "excluded": EXCLUDED}, indent=2, ensure_ascii=False),
    encoding="utf-8")

n_ti = sum(1 for v in mapping.values() if v.startswith("TI"))
n_it = sum(1 for v in mapping.values() if v.startswith("IT"))
print(f"Mappatura salvata: {len(mapping)} soggetti (TI: {n_ti}, IT: {n_it}); esclusi: {len(EXCLUDED)}")
for pid in sorted(mapping, key=sortkey):
    print(f"  {pid:4} -> {mapping[pid]}")
