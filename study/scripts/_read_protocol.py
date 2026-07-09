# -*- coding: utf-8 -*-
"""Estrae il testo dal protocollo .docx e il contenuto di test.xlsx."""
import zipfile
import re
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent

# --- docx: estrai testo dai paragrafi ---
docx = BASE / "protocols" / "Protocollo da tenere sul banco.docx"
with zipfile.ZipFile(docx) as z:
    xml = z.read("word/document.xml").decode("utf-8")
# paragrafi: separa per </w:p>, togli i tag
paras = []
for p in xml.split("</w:p>"):
    text = "".join(re.findall(r"<w:t[^>]*>([^<]*)</w:t>", p))
    if text.strip():
        paras.append(text.strip())
print("=== PROTOCOLLO ===")
for p in paras:
    print(p)

# --- test.xlsx ---
print("\n=== TEST.XLSX ===")
wb = openpyxl.load_workbook(BASE / "test.xlsx")
print("fogli:", wb.sheetnames)
for ws in wb.worksheets:
    print(f"\n--- foglio: {ws.title} ({ws.max_row} righe x {ws.max_column} colonne)")
    for row in ws.iter_rows(values_only=True):
        if any(v is not None for v in row):
            print(row)
