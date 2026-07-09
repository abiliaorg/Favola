# -*- coding: utf-8 -*-
"""Classifica ogni soggetto per ORDINE REALE di somministrazione (TI/IT) dai gaze.

- TI = prima registrazione text, seconda images; IT = l'inverso.
- L'ordine si ricava dai timestamp delle registrazioni, NON dal prefisso B/F
  (le postazioni a volte hanno chiamato bambini dell'altro gruppo o invertito l'ordine).
- La narratrice è la stessa per tutte le storie; bene/fra = postazione fisica.

Output: study/ordine_soggetti.xlsx + statistiche con i gruppi reali TI/IT.
"""
import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

STUDY = Path(__file__).resolve().parent.parent
GAZE = STUDY.parent / "data" / "02_gaze"

data = json.loads((STUDY / "chiave_derivata.json").read_text(encoding="utf-8"))
KEY = {k: v for k, v in data.items() if not k.startswith("_")}
EN2IT = {"fox": "volpe e boscaiolo", "carpet": "tappeto", "cats": "gatta", "yawn": "sbadiglio",
         "dolphin": "delfino", "panda": "panda", "bear": "orso", "eels": "anguille"}
FIRST_STORIES = {"volpe e boscaiolo", "gatta", "delfino", "anguille"}

# --- registrazioni gaze per soggetto ---
recs = {}
for folder in sorted(GAZE.iterdir()):
    if not folder.is_dir() or not folder.name.startswith("data_"):
        continue
    m = re.match(r"data_(bene|fra)(\d)", folder.name)
    station, sess = m.group(1), m.group(2)
    for f in sorted(folder.glob("*.json")):
        d = json.loads(f.read_text(encoding="utf-8"))
        pid = str(d.get("participantId", "")).strip("`").strip().upper()
        if not re.match(r"^[BF]\d+$", pid):
            continue
        st = EN2IT.get(d.get("story"))
        if not st:
            continue
        recs.setdefault(pid, []).append({
            "storia": st, "mod": d.get("typology"), "time": f.name[:15],
            "station": station, "sess": "mattina" if sess == "1" else "pomeriggio",
            "classe": str(d.get("class", "")),
        })

# --- punteggi test (per correlare con l'ordine reale) ---
wb = openpyxl.load_workbook(STUDY / "test.xlsx")
scores = {}
for r in list(wb["PROVE"].iter_rows(values_only=True))[1:]:
    if not r[0] or not r[1]:
        continue
    pid, storia = str(r[0]).strip().upper(), str(r[1]).strip()
    key = KEY[storia]
    given = [str(r[2+i]).strip().lower() if r[2+i] is not None else "" for i in range(len(key))]
    if all(a == "" for a in given):
        continue
    scores[(pid, storia)] = sum(1 for a, k in zip(given, key) if a == k) / len(key) * 100

# --- classificazione per soggetto ---
all_pids = sorted(set(recs) | {p for p, _ in scores},
                  key=lambda p: (p[0], int(re.sub(r"\D", "", p) or 0)))
subjects = []
for pid in all_pids:
    rr = sorted(recs.get(pid, []), key=lambda x: x["time"])
    cls = rr[0]["classe"] if rr else "?"
    if len(rr) == 2:
        seq = "".join("T" if x["mod"] == "text" else "I" for x in rr)
        order = seq if seq in ("TI", "IT") else f"ANOMALO ({seq})"
        fonte = "gaze"
    elif len(rr) == 1:
        known = rr[0]
        other = "images" if known["mod"] == "text" else "text"
        pos_known = 1 if known["storia"] in FIRST_STORIES else 2
        seq = ("T" if known["mod"] == "text" else "I") if pos_known == 1 else None
        if seq:
            order = (seq + ("I" if seq == "T" else "T")) + "?"
        else:
            order = (("T" if other == "text" else "I") + ("T" if known["mod"] == "text" else "I")) + "?"
        fonte = "presunto (1 gaze mancante)"
    else:
        order = "TI?" if pid[0] == "B" else "IT?"
        fonte = "presunto (nessun gaze, solo disegno)"
    stations = "/".join(dict.fromkeys(x["station"] for x in rr)) or "?"
    sess = rr[0]["sess"] if rr else "?"
    mismatch_station = ("B" if pid[0] == "B" else "F") not in {"bene": "B", "fra": "F"}.get(rr[0]["station"], "?") if rr else False
    # punteggi per posizione
    s = {}
    for x in rr:
        s[x["mod"]] = scores.get((pid, x["storia"]))
    subjects.append({
        "pid": pid, "classe": cls, "ordine": order, "fonte": fonte,
        "sessione": sess, "postazioni": stations,
        "dettaglio": " -> ".join(f"{x['time'][9:11]}:{x['time'][11:13]} {x['storia'][:10]}/{x['mod']}@{x['station']}" for x in rr),
        "score_text": s.get("text"), "score_images": s.get("images"),
        "n_test": sum(1 for st in KEY if (pid, st) in scores),
    })

# --- excel ---
out = openpyxl.Workbook()
ws = out.active
ws.title = "Ordine soggetti"
hdr = ["ID", "Classe", "Ordine reale", "Fonte", "Sessione", "Postazioni",
       "Sequenza (ora storia/mod@postazione)", "Score text %", "Score images %", "N test"]
ws.append(hdr)
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4472C4")
warn = PatternFill("solid", fgColor="FFEB9C")
for s in subjects:
    ws.append([s["pid"], s["classe"], s["ordine"], s["fonte"], s["sessione"], s["postazioni"],
               s["dettaglio"], s["score_text"], s["score_images"], s["n_test"]])
    if "?" in s["ordine"] or "ANOMALO" in s["ordine"]:
        for c in ws[ws.max_row]:
            c.fill = warn
for col, w in zip("ABCDEFGHIJ", [7, 8, 14, 28, 11, 11, 62, 12, 13, 8]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"
out.save(STUDY / "ordine_soggetti.xlsx")

def mean(xs): return sum(xs) / len(xs) if xs else float("nan")
def sd(xs):
    if len(xs) < 2: return float("nan")
    m = mean(xs)
    return (sum((x - m) ** 2 for x in xs) / (len(xs) - 1)) ** 0.5
def tstat(xs): return mean(xs) / (sd(xs) / len(xs) ** 0.5) if len(xs) > 1 and sd(xs) > 0 else float("nan")
def t2(a, b):
    va, vb = sd(a) ** 2 / len(a), sd(b) ** 2 / len(b)
    return (mean(a) - mean(b)) / (va + vb) ** 0.5

print(f"Salvato: {STUDY / 'ordine_soggetti.xlsx'}  ({len(subjects)} soggetti)")

print("\n=== Classificazione ordine reale ===")
for o in sorted({s["ordine"] for s in subjects}):
    ids = [s["pid"] for s in subjects if s["ordine"] == o]
    print(f"  {o:14} n={len(ids):2}  {', '.join(ids)}")

print("\n=== Prefisso vs ordine reale (solo gaze certi) ===")
for s in subjects:
    if s["ordine"] in ("TI", "IT"):
        expected = "TI" if s["pid"][0] == "B" else "IT"
        if s["ordine"] != expected:
            print(f"  {s['pid']}: prefisso {s['pid'][0]} ma ordine reale {s['ordine']}  ({s['dettaglio']})")

print("\n=== Postazioni incrociate (bambino registrato su entrambe o sulla postazione 'altrui') ===")
for s in subjects:
    if "/" in s["postazioni"]:
        print(f"  {s['pid']}: {s['postazioni']}  ({s['dettaglio']})")

# --- statistiche con i gruppi REALI ---
def diffs_for(order):
    out = []
    for s in subjects:
        if s["ordine"] == order and s["score_text"] is not None and s["score_images"] is not None:
            out.append(s["score_images"] - s["score_text"])
    return out

dti = diffs_for("TI")   # images-text = S2-S1
dit = diffs_for("IT")   # images-text = S1-S2
print("\n=== 2bis) Ordine reale: TI vs IT (media dei due test per bambino) ===")
sti = [mean([s["score_text"], s["score_images"]]) for s in subjects
       if s["ordine"] == "TI" and s["score_text"] is not None and s["score_images"] is not None]
sit = [mean([s["score_text"], s["score_images"]]) for s in subjects
       if s["ordine"] == "IT" and s["score_text"] is not None and s["score_images"] is not None]
print(f"  TI: n={len(sti):2}  media={mean(sti):5.1f}%  sd={sd(sti):4.1f}")
print(f"  IT: n={len(sit):2}  media={mean(sit):5.1f}%  sd={sd(sit):4.1f}")
print(f"  diff (IT-TI)={mean(sit)-mean(sti):+.1f} pt  t(Welch)={t2(sit, sti):.2f}")

print("\n=== 3bis) Decomposizione con gruppi reali ===")
alld = dti + dit
print(f"  appaiato complessivo (images-text): n={len(alld)}  {mean(alld):+.1f} pt  t={tstat(alld):.2f}")
print(f"  nei TI (=S2-S1): n={len(dti):2}  {mean(dti):+.1f} pt")
print(f"  negli IT (=S1-S2): n={len(dit):2}  {mean(dit):+.1f} pt")
m = (mean(dti) + mean(dit)) / 2
p = (mean(dti) - mean(dit)) / 2
se = ((sd(dti) ** 2 / len(dti) + sd(dit) ** 2 / len(dit)) ** 0.5) / 2
print(f"  -> effetto MODALITÀ  m={m:+.1f} pt (t≈{m/se:.2f})")
print(f"  -> effetto POSIZIONE/STORIA (S2-S1) p={p:+.1f} pt (t≈{p/se:.2f})")

print("\n=== Bilanciamento classe x ordine reale (coppie certe) ===")
classes = sorted({s["classe"] for s in subjects})
print(f"  {'classe':8} {'TI':>4} {'IT':>4}")
for cl in classes:
    nti = sum(1 for s in subjects if s["classe"] == cl and s["ordine"] == "TI")
    nit = sum(1 for s in subjects if s["classe"] == cl and s["ordine"] == "IT")
    print(f"  {cl:8} {nti:4} {nit:4}")
