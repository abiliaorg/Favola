# -*- coding: utf-8 -*-
"""Corregge il pid '`F4' e rigenera 'registrazioni_gaze tutte.xlsx' (senza test/prova)."""
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent.parent / "data" / "02_gaze"

# --- fix backtick in F4 fox images ---
f4 = BASE / "data_fra1" / "20260706_100520_F4_2_fox_images.json"
raw = f4.read_text(encoding="utf-8")
if '"participantId":"`F4"' in raw:
    f4.write_text(raw.replace('"participantId":"`F4"', '"participantId":"F4"'), encoding="utf-8")
    print("Corretto participantId '`F4' -> 'F4' in", f4.name)
else:
    print("Backtick non trovato in", f4.name, "- nessuna modifica")

# --- rigenera 'tutte' (esclusi test/prova/demo) ---
TEST_IDS = re.compile(r"^(test\d*|test_audio|testaudio|prova|tommy|giulia|vittorio)$", re.IGNORECASE)
rows = []
for folder in sorted(d for d in BASE.iterdir() if d.is_dir() and d.name.startswith("data_")):
    pc = "Benedetta" if folder.name.startswith("data_bene") else "Francesca"
    for f in sorted(folder.glob("*.json")):
        data = json.loads(f.read_text(encoding="utf-8"))
        pid = str(data.get("participantId", ""))
        if TEST_IDS.match(pid):
            continue
        samples = data.get("samples") or []
        m = re.match(r"(\d{4})(\d{2})(\d{2})_(\d{2})(\d{2})(\d{2})_", f.name)
        date_str = f"{m.group(3)}/{m.group(2)}/{m.group(1)}" if m else ""
        time_str = f"{m.group(4)}:{m.group(5)}:{m.group(6)}" if m else ""
        rows.append([pc, folder.name, f.name, date_str, time_str, pid,
                     data.get("class", ""), data.get("story", ""), data.get("typology", ""),
                     round(samples[-1]["t"], 1) if samples else 0, len(samples),
                     "sì" if data.get("calibration") else "no", None])

rows.sort(key=lambda r: (r[3], r[4]))

wb = Workbook()
ws = wb.active
ws.title = "Registrazioni"
headers = ["PC", "Cartella", "File", "Data", "Ora", "Partecipante", "Classe", "Storia",
           "Modalità", "Durata (s)", "N. campioni", "Calibrazione", "Test"]
ws.append(headers)
header_fill = PatternFill("solid", fgColor="4472C4")
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
for r in rows:
    ws.append(r)
widths = [12, 13, 46, 11, 10, 13, 8, 10, 10, 11, 12, 13, 6]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
ws.freeze_panes = "A2"

out = BASE / "registrazioni_gaze.xlsx"
try:
    wb.save(out)
except PermissionError:
    out = out.with_name(out.stem + "_new.xlsx")
    wb.save(out)
    print("ATTENZIONE: file aperto in Excel, salvato come", out.name)
print(f"Salvato {out.name}: {len(rows)} registrazioni")
