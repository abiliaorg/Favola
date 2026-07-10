# -*- coding: utf-8 -*-
"""Finestre temporali dell'informazione rilevante per ogni domanda dei test.

Per ogni domanda è definita un'ancora testuale: il periodo del transcript che
contiene l'informazione necessaria a rispondere. L'ancora viene localizzata
nella sequenza di parole (non-interim) di ciascuna variante (text/images) con
matching fuzzy (tollera i refusi del riconoscimento vocale), e convertita in
finestra temporale usando il videoTime di prima comparsa delle parole.

Domande "globali" (argomento del brano, tipo di testo, definizioni) non hanno
un momento puntuale e sono marcate type=globale.

Output:
  study/question_windows.json  — per storia/domanda/variante: t0, t1, parole
  study/question_windows.xlsx  — tabella di revisione umana
"""
import json
import re
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

STUDY = Path(__file__).resolve().parent.parent
RECORD = STUDY.parent / "data" / "01_record"

PAD_BEFORE = 1.0   # secondi prima della prima parola dell'ancora
PAD_AFTER = 1.5    # secondi dopo l'ultima parola (tempo di elaborazione)

IT2EN = {"volpe e boscaiolo": "fox", "tappeto": "carpet", "gatta": "cats",
         "sbadiglio": "yawn", "delfino": "dolphin", "panda": "panda",
         "orso": "bear", "anguille": "eels"}
STORY_CLASSES = {"volpe e boscaiolo": ["1", "2"], "tappeto": ["1", "2"],
                 "gatta": ["3"], "sbadiglio": ["3"], "delfino": ["4"],
                 "panda": ["4"], "anguille": ["5"], "orso": ["5"]}

# Ancore: periodo del transcript (variante text) che contiene l'informazione.
# None = domanda globale (nessun momento puntuale). "nota" opzionale.
ANCHORS = {
"volpe e boscaiolo": {
  1: "per pietà nascondimi chiese al taglialegna",
  2: "Pochi istanti dopo arrivarono i cacciatori e domandarono al taglialegna",
  3: "giunse alla Capanna di un taglialegna",
  4: "giunse alla Capanna di un taglialegna per pietà nascondimi chiese al taglialegna",
  5: "entra nella mia capanna disse L'uomo nasconditi in quel Cantuccio",
  6: "e intanto con la mano faceva segno che sì la l'aveva volpe vista e anzi era lì in quel calduccio della Capanna",
  7: "Ma i cacciatori mandarono solo alle sue parole e non ai gesti e si allontanarono a grandi passi",
  8: "poi uscì anch'essa e si avviò dalla parte opposta l'uomo fece l'offeso Ma come ti ho salvato e tu te ne vai senza nemmeno dire grazie Ti ringrazierei rispose la volpe",
  9: None,  # definizione di "boscaiolo": globale
  10: "l'uomo fece l'offeso Ma come ti ho salvato e tu te ne vai senza nemmeno dire grazie",
},
"tappeto": {
  1: "non ha più visto in mezzo alla folla né il suo papà né la sua mamma",
  2: "è andato a guardare una scimmia",
  3: "Allora si mise a piangere perché piangi gli ha chiesto allora un vecchio che vendeva tappeti Il bambino ha risposto che si era perduto",
  4: "il tappeto che era magico si è alzato sopra il mercato",   # nota: informazione indiretta
  5: "perché piangi gli ha chiesto allora un vecchio che vendeva tappeti",
  6: "tornare a casa siediti su questo tappeto è un tappeto speciale che non vendo a nessuno",
  7: "si è alzato sopra il mercato sopra la città e i campi",
  8: "ed è atterrato proprio davanti alla sua casa",
  9: "il bambino è sceso felice",
  10: "il tappeto tornava dal suo padrone",
},
"gatta": {
  1: "nel giardino della mia casa di campagna",
  2: "mentre cadevano fiocchi di neve nel giardino della mia casa di campagna apparve una gatta",
  3: "apparve una gatta bianca e nera che zoppicava",
  4: "ebbi con passione di quella povera gatta e Le un po' di latte caldo e un pezzetto di carne di pollo",
  5: "Io rimasi a guardare la neve che imbiancava il prato",
  6: "non passarono neanche tre minuti che la gatta zoppicando ricomparve",
  7: "ricomparve da dietro l'angolo della casa con gli occhi mestie",
  8: "poi tornò indietro come per cercare qualcosa ricomparve con tre piccoli gattini infreddoliti uno dietro l'altro gli portò vicino a me",
  9: "ricomparve con tre piccoli gattini infreddoliti",
  10: "avevano freddo e fame e diede a loro il latte e la carne",
  11: "mentre la mamma felice li leccava per pulirli",
  12: "misi tutta la famiglia in una cassetta collana e la portai al riparo nei ripostiglio della legna",
},
"sbadiglio": {
  1: "portò a passeggio la sorellina nel passeggino",
  2: "La donna del negozio di cappelli La vide e subito sbadigliò anche lei",
  3: "La sorellina era stanca e sbadigliò",
  4: "il manovratore vedendo tutte quelle bocche aperte cominciò a sbadigliare e sbadigliare e non riusciva a ripartire",
  5: "il vigile voleva suonare il fischietto perché tutti si decidessero a rimettersi in moto ma non riuscì a fischiare perché aveva voglia di sbadigliare",
  6: "un camionista che voleva sapere perché il tram stesse fermo per tanto tempo si sporse dal finestrino e subito lo sbadiglio colse anche lui",
  7: "e il giornalaio il ciclista tutti si misero a sbadigliare",
  8: "anche lo spazzacamino sul tetto",
  9: "tutta la gente tutti i cani e i gatti della città si a sbadigliare anche lo spazzacamino sul tetto e persino I lombrichi nella terra",
  10: "per fortuna fu presto sera e tutti andarono a dormire di buon'ora",
},
"delfino": {
  1: "era un giorno bellissimo di fine luglio",
  2: "il mare mi accarezzava E cantava un canto lieve tenue come una linea da bambini",
  3: "di tanto in tanto mi lasciavo trasportare inerte dalle onde il mare mi accarezzava",
  4: "vidi lontano un sommovimento di acque e un triangolo nero avanzare",
  5: "tanto Fra poco sarò fra le sue zanne",
  6: "decisi di tornare indietro e come nulla fosse vi diressi verso la lontana Riva",
  7: "Se era uno squalo è inutile affannarsi tanto Fra poco sarò fra le sue zanne se non lo è tanto meglio",
  8: "non era uno squalo era un grosso delfino e aveva voglia di giocare",
  9: "si mise a fare intorno a me giri vorticosi Poi salti fuori dalle onde Così che vedevo ne tutto il grande e bellissimo corpo poi scomparve passo un po' di tempo",
  10: "si era posto con la schiena proprio sotto di me",
  11: "alla fine stanco di saltare e di giocare il delfino scomparve",
  12: "e io lentamente tornai arriva",
  13: "mi dimenticai delle bandiere rosse e presi in largo",
  14: None,  # titolo alternativo: globale
},
"panda": {
  1: "è un buffo timido simpatico orso di abitudini ritirate",
  2: "molto bravo ad arrampicarsi sugli alberi grazie alle sue dimensioni ridotte",
  3: "della più grande Organizzazione Internazionale per la conservazione degli animali il fondo mondiale per la natura",
  4: "tranne che allo zoo di Pechino non si è mai riusciti a far riprodurre in cattività",
  5: "non si è mai riusciti a far riprodurre in cattività Questo graziosissimo animale",
  6: "un po' per la sua rarità un po' per il suo aspetto grazioso",
  7: "con gli orsi ha solo una lontana parentina perché appartiene alla stessa famiglia dei procioni",
  8: "il panda si nutre esclusivamente di germogli di bambù",
  9: "Infatti ne mangia all'incirca 12 kg al giorno",
  10: "che mangia ingrossa quantità dato che si tratta di un alimento poco nutriente",
  11: "abitatore delle foreste di bambù di alta montagna del Tibet orientale e della Cina meridionale",
  12: "una specie di sesto udito gli permette di afferrare anche i pezzetti più piccoli",
  13: None,  # argomento principale: globale
  14: None,  # tipo di brano: globale
},
"orso": {
  1: "compie lunghe e migrazioni che lo portano anche migliaia di chilometri lontano dal punto di partenza Alla ricerca delle zone dove le foche Sono più numerosi",
  2: "deve nutrirsi moltissimo per accumulare enormi quantità di grasso solo così potrà superare il lungo e rigido inverno",
  3: "il lungo e rigido inverno durante il quale è praticamente impossibile trovare qualche preda",
  4: "durante il quale è praticamente impossibile trovare qualche preda",
  5: "la parte posteriore grossa e massiccia contrasta con quella anteriore meno tozza",
  6: "le zampe sono piuttosto grandi adatte ad una vita sul ghiaccio e sulla neve",
  7: "le dita sono unite da una membrana che gli facilita il nuoto",
  8: "robusti e corti artigli rendono micidiali le poderosi zampe",
  9: "il suo olfatto straordinario gli Segnala la presenza di prede anche apparecchi km di distanza",
  10: "il grosso plantigrado si avvicina strisciando carponi giunto a distanza ravvicinata uccide la preda con una zampata",
  11: "a volte si spinge con in acqua e a nuoto si dirige verso il blocco di ghiaccio sul quale ha adocchiato un pesce",
  12: "il blocco di ghiaccio sul quale ha adocchiato un pesce emerge Quindi con un grande balzo",
  13: "a volte si spinge con in acqua e a nuoto si dirige verso il blocco di ghiaccio",
  14: "attende vicino ad un foro del ghiaccio che una foca emerga per respirare quando la vede subito la afferra e la uccide",
},
"anguille": {
  1: "quando si avvicina il tempo di deporre le uova l'anguilla sente un grande bisogno di andare lungo la corrente di un fiume",
  2: "passa per lo stretto di Gibilterra ed entra nell'Oceano Atlantico che attraversa quasi completamente",
  3: "quando è giunta in una zona piena di alghe chiamata mar dei Sargassi",
  4: "essa depone le uova e poi muore scomparendo nelle profondità marine",
  5: "quando le uova si schiudono le piccole anguille lunghe solo qualche centimetro iniziano il viaggio di ritorno",
  6: "quelli nati da anguille americane vanno verso l'America",
  7: "il viaggio per queste ultime È lunghissimo dura tre anni",
  8: "le piccole anguille lunghe solo qualche centimetro",
  9: "in una zona piena di alghe chiamata mar dei Sargassi essa depone le uova",
  10: "essa depone le uova e poi muore scomparendo nelle profondità marine",
},
}

NOTES = {
  ("tappeto", 4): "informazione indiretta: il mercato è esplicito solo al decollo del tappeto",
  ("orso", 12): "la narratrice dice 'pinnipede' ma il riconoscimento ha scritto 'pesce' (ed è ciò che la caption mostrava; in images era un'immagine): finestra ancorata alla frase riconosciuta; risposta in parte inferenziale",
  ("delfino", 13): "informazione all'inizio del brano, risposta inferenziale",
}

def norm_tokens(text):
    text = text.lower()
    text = re.sub(r"[’']", " ", text)
    text = re.sub(r"[^\wàèéìòù]+", " ", text)
    return [t for t in text.split() if t]

def load_words(cls, story_en, typ):
    """Sequenza di parole non-interim con tempo di prima comparsa."""
    d = json.loads((RECORD / f"{cls}_{story_en}_{typ}.json").read_text(encoding="utf-8"))
    meta = d.get("wordMetaById") or {}
    first_t = {}
    for snap in d.get("textUpdateSnapshots") or []:
        for e in snap.get("entries") or []:
            wid = str(e["wordAutoIncrementalId"])
            t = float(e.get("videoTime", snap.get("videoTime", 0)))
            if wid not in first_t or t < first_t[wid]:
                first_t[wid] = t
    words = []
    for wid in sorted(meta, key=int):
        m = meta[wid]
        if m.get("interim"):
            continue
        toks = norm_tokens(m.get("word", ""))
        if not toks:
            continue
        words.append({"id": wid, "word": m.get("word", ""), "toks": toks,
                      "t": first_t.get(wid)})
    return words

def find_window(words, anchor_toks):
    """Migliore finestra contigua di parole che matcha l'ancora (fuzzy)."""
    seq = []          # (word_index, token) espansi
    for i, w in enumerate(words):
        for tok in w["toks"]:
            seq.append((i, tok))
    toks = [t for _, t in seq]
    target = " ".join(anchor_toks)
    n, m = len(toks), len(anchor_toks)
    best = (0.0, None, None)
    for delta in (0, -1, 1, -2, 2):
        L = m + delta
        if L < 1 or L > n:
            continue
        for start in range(0, n - L + 1):
            cand = " ".join(toks[start:start + L])
            r = SequenceMatcher(None, cand, target).ratio()
            if r > best[0]:
                best = (r, start, start + L - 1)
    ratio, s, e = best
    if s is None:
        return None
    wi0, wi1 = seq[s][0], seq[e][0]
    return {"ratio": round(ratio, 3), "wi0": wi0, "wi1": wi1}

results = {}
review_rows = []
warnings = []
for story_it, anchors in ANCHORS.items():
    story_en = IT2EN[story_it]
    results[story_it] = {}
    variants = {}
    for cls in STORY_CLASSES[story_it]:
        for typ in ("text", "images"):
            key = f"{cls}_{typ}"
            try:
                variants[key] = load_words(cls, story_en, typ)
            except FileNotFoundError:
                pass
    for q, anchor in sorted(anchors.items()):
        entry = {"type": "puntuale" if anchor else "globale",
                 "anchor": anchor, "windows": {}}
        note = NOTES.get((story_it, q))
        if note:
            entry["note"] = note
        if anchor:
            atoks = norm_tokens(anchor)
            for key, words in variants.items():
                w = find_window(words, atoks)
                if not w:
                    warnings.append(f"{story_it} Q{q} {key}: nessun match")
                    continue
                span = words[w["wi0"]:w["wi1"] + 1]
                times = [x["t"] for x in span if x["t"] is not None]
                if not times:
                    warnings.append(f"{story_it} Q{q} {key}: parole senza tempo")
                    continue
                t0, t1 = min(times), max(times)
                matched = " ".join(x["word"] for x in span)
                if w["ratio"] < 0.6:
                    warnings.append(f"{story_it} Q{q} {key}: match debole ({w['ratio']}) -> '{matched[:60]}'")
                entry["windows"][key] = {
                    "t0": round(max(0.0, t0 - PAD_BEFORE), 2),
                    "t1": round(t1 + PAD_AFTER, 2),
                    "t_first_word": round(t0, 2), "t_last_word": round(t1, 2),
                    "word_ids": [x["id"] for x in span],
                    "matched_text": matched, "match_ratio": w["ratio"],
                }
                review_rows.append([story_it, q, key, entry["type"],
                                    round(t0, 1), round(t1, 1), w["ratio"],
                                    anchor[:70], matched[:70]])
        else:
            review_rows.append([story_it, q, "-", "globale", None, None, None,
                                "(nessun momento puntuale)", ""])
        results[story_it][str(q)] = entry

meta = {"_pad_before_s": PAD_BEFORE, "_pad_after_s": PAD_AFTER,
        "_nota": "t0/t1 = finestra con padding; t_first/last_word = comparsa parole ancora. "
                 "Chiavi finestra = {classe}_{tipologia} (le due riprese hanno tempi diversi)."}
out = {"_meta": meta, **results}
(STUDY / "question_windows.json").write_text(
    json.dumps(out, indent=1, ensure_ascii=False), encoding="utf-8")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Finestre"
ws.append(["Storia", "Domanda", "Variante", "Tipo", "t prima parola", "t ultima parola",
           "Match ratio", "Ancora (inizio)", "Testo trovato (inizio)"])
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4472C4")
warn_fill = PatternFill("solid", fgColor="FFEB9C")
for r in review_rows:
    ws.append(r)
    if r[6] is not None and r[6] < 0.6:
        for c in ws[ws.max_row]:
            c.fill = warn_fill
for col, w in zip("ABCDEFGHI", [18, 9, 10, 9, 13, 13, 11, 52, 52]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:I{ws.max_row}"
wb.save(STUDY / "question_windows.xlsx")

n_punt = sum(1 for s in results.values() for e in s.values() if e["type"] == "puntuale")
n_glob = sum(1 for s in results.values() for e in s.values() if e["type"] == "globale")
n_win = sum(len(e["windows"]) for s in results.values() for e in s.values())
print(f"Domande: {n_punt} puntuali + {n_glob} globali; finestre calcolate: {n_win}")
print(f"Salvati: question_windows.json, question_windows.xlsx")
if warnings:
    print("\nAVVISI:")
    for w in warnings:
        print("  -", w)
