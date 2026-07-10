# -*- coding: utf-8 -*-
"""Motore AOI: porting fedele di web/analysis/assets/index.js in Python.

Per ogni registrazione gaze dei 48 soggetti validi calcola, campione per
campione, la categoria guardata (caption word/image/band, face e sue parti,
none) con le stesse regole del tool web:
  - calibrazione a 5 punti per asse (fit lineare misurato->target), ATTIVA
  - gaze -> frazione frame con object-fit:contain (viewport sessione vs video)
  - box -> frazione con denomH = recordVp.width * vidH/vidW
  - trasformazione per categoria (scala attorno al centro + offset) da
    study/aoi_params.json
  - priorità: word/image specifica > parte volto > volto intero > caption band > none

Video verificati tutti 1920x1080 (ffprobe, 9/7/2026).

Output: study/aoi_results.xlsx + study/aoi_results.json (aggregati per
registrazione, inclusi i conteggi per parola per il lavoro puntuale futuro).
"""
import json
import re
from bisect import bisect_right
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

STUDY = Path(__file__).resolve().parent.parent
ROOT = STUDY.parent
GAZE = ROOT / "data" / "02_gaze"
RECORD = ROOT / "data" / "01_record"

VID_W, VID_H = 1920, 1080  # dimensione intrinseca di tutti i video (verificata)

params = json.loads((STUDY / "aoi_params.json").read_text(encoding="utf-8"))
ALPHA = params["alphaByCat"]
USE_CAL = bool(params.get("calibration", True))
GAZE_DY = float(params.get("gazePointerDyPx", 0))

_map = json.loads((STUDY / "mappatura_soggetti.json").read_text(encoding="utf-8"))
MAPPING = _map["mapping"]

EN2IT = {"fox": "volpe e boscaiolo", "carpet": "tappeto", "cats": "gatta", "yawn": "sbadiglio",
         "dolphin": "delfino", "panda": "panda", "bear": "orso", "eels": "anguille"}

# ---------- calibrazione (fitAxisLinear + computeGazeCalibration) ----------

def fit_axis_linear(ms, ts):
    n = len(ms)
    if n == 0:
        return (1.0, 0.0)
    mm = sum(ms) / n
    mt = sum(ts) / n
    num = sum((m - mm) * (t - mt) for m, t in zip(ms, ts))
    den = sum((m - mm) ** 2 for m in ms)
    if n < 2 or den < 1e-6:
        return (1.0, mt - mm)
    a = num / den
    return (a, mt - a * mm)

def compute_calibration(cal):
    if not cal or not isinstance(cal.get("points"), list):
        return None
    pts = [p for p in cal["points"] if p and p.get("measured") and p.get("target")]
    if len(pts) == 1:
        p = pts[0]
        return ((1.0, p["target"]["x"] - p["measured"]["x"]),
                (1.0, p["target"]["y"] - p["measured"]["y"]))
    if len(pts) < 2:
        return None
    cx = fit_axis_linear([p["measured"]["x"] for p in pts], [p["target"]["x"] for p in pts])
    cy = fit_axis_linear([p["measured"]["y"] for p in pts], [p["target"]["y"] for p in pts])
    return (cx, cy)

# ---------- geometria ----------

def gaze_to_fraction(gx, gy, sess_w, sess_h, cal):
    s = min(sess_w / VID_W, sess_h / VID_H)
    sw, sh = VID_W * s, VID_H * s
    if sw <= 0 or sh <= 0:
        return None
    if USE_CAL and cal:
        (ax, bx), (ay, by) = cal
        gx = ax * gx + bx
        gy = ay * gy + by
    return (gx / sw, (gy + GAZE_DY) / sh)

def box_to_fraction(box, rec_w, rec_h):
    if not box or not rec_w or not rec_h:
        return None
    denom_h = rec_w * VID_H / VID_W
    return (box["x"] / rec_w, box["y"] / denom_h, box["w"] / rec_w, box["h"] / denom_h)

def apply_alpha(frac, cat):
    a = ALPHA.get(cat, {"w": 1.0, "h": 1.0, "dx": 0.0, "dy": 0.0})
    x, y, w, h = frac
    if a["w"] != 1 or a["h"] != 1:
        cx, cy = x + w / 2, y + h / 2
        w, h = w * a["w"], h * a["h"]
        x, y = cx - w / 2, cy - h / 2
    return (x + a["dx"], y + a["dy"], w, h)

def in_box(px, py, b):
    return b[0] <= px <= b[0] + b[2] and b[1] <= py <= b[1] + b[3]

# ---------- tracking per storia (cache) ----------

_tracking_cache = {}

def load_tracking(cls, story, typ):
    key = (cls, story, typ)
    if key in _tracking_cache:
        return _tracking_cache[key]
    d = json.loads((RECORD / f"{cls}_{story}_{typ}.json").read_text(encoding="utf-8"))
    vp = d.get("viewport") or {"width": 1920, "height": 1024}
    words = sorted(d.get("textUpdateSnapshots") or [], key=lambda s: float(s["videoTime"]))
    faces = sorted(d.get("faceTrackingSnapshots") or [], key=lambda s: float(s["videoTime"]))
    meta = d.get("wordMetaById") or {}
    wtimes = [float(s["videoTime"]) for s in words]
    ftimes = [float(s["videoTime"]) for s in faces]

    # pre-calcolo: per ogni snapshot parole, i box (frazione + alpha) delle entry
    word_boxes = []
    for snap in words:
        svp = snap.get("viewport") or vp
        entries = []
        for e in snap.get("entries") or []:
            frac = box_to_fraction(e.get("location"), svp.get("width"), svp.get("height"))
            if not frac:
                continue
            m = meta.get(str(e["wordAutoIncrementalId"])) or {}
            sub = "image" if m.get("image") else "word"
            acat = "image" if sub == "image" else "text"
            entries.append((apply_alpha(frac, acat), sub, str(e["wordAutoIncrementalId"]), m.get("word", "")))
        word_boxes.append(entries)

    face_boxes = []
    for snap in faces:
        svp = snap.get("viewport") or vp
        boxes = snap.get("boxes") or {}
        parts = []
        for part in ("face", "mouth", "nose", "eyeLeft", "eyeRight"):
            frac = box_to_fraction(boxes.get(part), svp.get("width"), svp.get("height"))
            if frac:
                parts.append((apply_alpha(frac, part), part))
        face_boxes.append(parts)

    # caption band: y minimo su tutte le entry (senza alpha)
    top = None
    for snap in words:
        svp = snap.get("viewport") or vp
        for e in snap.get("entries") or []:
            frac = box_to_fraction(e.get("location"), svp.get("width"), svp.get("height"))
            if frac and (top is None or frac[1] < top):
                top = frac[1]
    band = (0.0, top, 1.0, max(0.0, 1.0 - top)) if top is not None else None

    tr = {"wtimes": wtimes, "word_boxes": word_boxes,
          "ftimes": ftimes, "face_boxes": face_boxes,
          "band": band, "meta": meta}
    _tracking_cache[key] = tr
    return tr

def at_or_before(times, t):
    i = bisect_right(times, t) - 1
    return i  # -1 se nessuno

# ---------- categorizzazione di un campione ----------

def categorize(gf, t, tr):
    """Ritorna (primary, sub, word_id) con le stesse priorità del tool web."""
    if gf is None:
        return ("none", "none", None)
    px, py = gf
    hits = []

    wi = at_or_before(tr["wtimes"], t)
    # NB: il tool web usa sempre arr[i] con i>=0 (clamp a 0): replichiamo il clamp
    if tr["word_boxes"]:
        for box, sub, wid, _w in tr["word_boxes"][max(wi, 0)]:
            if in_box(px, py, box):
                hits.append(("caption", sub, wid))
    if not hits and tr["band"] and in_box(px, py, tr["band"]):
        hits.append(("caption", "none", None))

    if tr["face_boxes"]:
        fi = at_or_before(tr["ftimes"], t)
        for box, part in tr["face_boxes"][max(fi, 0)]:
            if in_box(px, py, box):
                hits.append(("face", part, None))

    if not hits:
        return ("none", "none", None)
    chosen = (next((h for h in hits if h[0] == "caption" and h[2] is not None), None)
              or next((h for h in hits if h[0] == "face" and h[1] != "face"), None)
              or next((h for h in hits if h[0] == "face"), None)
              or hits[0])
    return chosen

# ---------- elaborazione delle registrazioni ----------

results = []
for folder in sorted(GAZE.iterdir()):
    if not folder.is_dir() or not folder.name.startswith("data_"):
        continue
    for f in sorted(folder.glob("*.json")):
        d = json.loads(f.read_text(encoding="utf-8"))
        pid = str(d.get("participantId", "")).strip("`").strip().upper()
        if pid not in MAPPING:
            continue
        story = d.get("story")
        if story not in EN2IT:
            continue
        cls, typ = str(d.get("class")), d.get("typology")
        tr = load_tracking(cls, story, typ)
        vp = d.get("viewport") or {}
        sess_w, sess_h = vp.get("width"), vp.get("height")
        cal = compute_calibration(d.get("calibration"))
        samples = d.get("samples") or []

        counts = {"caption": 0, "face": 0, "none": 0}
        csub = {"word": 0, "image": 0, "none": 0}
        fsub = {"face": 0, "mouth": 0, "nose": 0, "eyeLeft": 0, "eyeRight": 0}
        words = {}
        for s in samples:
            g = s.get("gaze")
            gf = gaze_to_fraction(g["x"], g["y"], sess_w, sess_h, cal) if g else None
            cat, sub, wid = categorize(gf, float(s["t"]), tr)
            counts[cat] += 1
            if cat == "caption":
                csub[sub] += 1
                if wid is not None:
                    words[wid] = words.get(wid, 0) + 1
            elif cat == "face":
                fsub[sub] += 1

        n = len(samples)
        dur = float(samples[-1]["t"]) if samples else 0.0
        pct = lambda v: round(v / n * 100, 1) if n else None
        top_words = sorted(words.items(), key=lambda kv: -kv[1])[:10]
        results.append({
            "pid": pid, "newid": MAPPING[pid], "grp": MAPPING[pid][:2],
            "classe": cls, "storia": EN2IT[story], "story_en": story, "mod": typ,
            "file": f.name, "n": n, "dur": round(dur, 1),
            "hz": round(n / dur, 1) if dur else None,
            "cal_pts": len((d.get("calibration") or {}).get("points") or []),
            "pct_caption": pct(counts["caption"]), "pct_face": pct(counts["face"]),
            "pct_none": pct(counts["none"]),
            "pct_word": pct(csub["word"]), "pct_image": pct(csub["image"]),
            "pct_band": pct(csub["none"]),
            "pct_mouth": pct(fsub["mouth"]), "pct_nose": pct(fsub["nose"]),
            "pct_eyes": pct(fsub["eyeLeft"] + fsub["eyeRight"]),
            "pct_face_only": pct(fsub["face"]),
            "top_words": [{"id": wid, "word": (tr["meta"].get(wid) or {}).get("word", ""),
                           "image": bool((tr["meta"].get(wid) or {}).get("image")),
                           "count": c} for wid, c in top_words],
            "word_counts": words,
        })

# ---------- output ----------

(STUDY / "aoi_results.json").write_text(
    json.dumps(results, indent=1, ensure_ascii=False), encoding="utf-8")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "AOI"
hdr = ["ID nuovo", "ID orig.", "Gruppo", "Classe", "Storia", "Modalità", "N campioni",
       "Durata (s)", "Hz", "Cal pts", "% caption", "% face", "% none",
       "% word", "% image", "% band", "% mouth", "% nose", "% eyes", "% face-only",
       "Top parole (id:conteggio)"]
ws.append(hdr)
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4472C4")
for r in sorted(results, key=lambda r: (r["grp"], int(r["newid"][2:]), r["storia"])):
    ws.append([r["newid"], r["pid"], r["grp"], r["classe"], r["storia"], r["mod"],
               r["n"], r["dur"], r["hz"], r["cal_pts"],
               r["pct_caption"], r["pct_face"], r["pct_none"],
               r["pct_word"], r["pct_image"], r["pct_band"],
               r["pct_mouth"], r["pct_nose"], r["pct_eyes"], r["pct_face_only"],
               ", ".join(f"{w['word'] or '#'+w['id']}:{w['count']}" for w in r["top_words"][:5])])
widths = [9, 8, 8, 7, 18, 10, 11, 10, 6, 8, 10, 8, 8, 8, 9, 8, 9, 8, 8, 11, 46]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:U{ws.max_row}"
wb.save(STUDY / "aoi_results.xlsx")

print(f"Elaborate {len(results)} registrazioni "
      f"({len({r['newid'] for r in results})} soggetti)")
print(f"Salvati: aoi_results.xlsx, aoi_results.json")

def mean(xs):
    xs = [x for x in xs if x is not None]
    return sum(xs) / len(xs) if xs else float("nan")

print("\n=== Medie AOI per modalità ===")
for mod in ("text", "images"):
    rs = [r for r in results if r["mod"] == mod]
    print(f"  {mod:6} (n={len(rs)}): caption={mean([r['pct_caption'] for r in rs]):.1f}%"
          f"  face={mean([r['pct_face'] for r in rs]):.1f}%"
          f"  none={mean([r['pct_none'] for r in rs]):.1f}%"
          f"  | word={mean([r['pct_word'] for r in rs]):.1f}%"
          f"  image={mean([r['pct_image'] for r in rs]):.1f}%"
          f"  band={mean([r['pct_band'] for r in rs]):.1f}%"
          f"  | mouth={mean([r['pct_mouth'] for r in rs]):.1f}%"
          f"  eyes={mean([r['pct_eyes'] for r in rs]):.1f}%")

low_q = [r for r in results if (r["hz"] or 0) < 5 or r["dur"] < 40]
if low_q:
    print("\n=== Registrazioni con qualità bassa (hz<5 o dur<40s) ===")
    for r in low_q:
        print(f"  {r['newid']} {r['storia']}/{r['mod']}: {r['n']} campioni, {r['dur']}s, {r['hz']} Hz")
